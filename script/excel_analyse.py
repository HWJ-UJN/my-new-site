# -*- coding: utf-8 -*-
"""
Excel 日记增量提取工具
- 支持多个 xlsx 文件
- 自动提取图片并映射到行号
- 增量更新数据，避免重复
- 生成统一的数据文件
"""
import openpyxl
import zipfile
import os
import shutil
import json
import sys
from datetime import datetime
from pathlib import Path

# 设置输出编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ==================== 配置区 ====================
# Excel 文件存放目录
EXCEL_DIR = r'D:\Code\d2l-zh\my-new-site\script'
# 图片输出目录
IMAGES_OUTPUT_DIR = r'D:\Code\d2l-zh\my-new-site\images\excel'
# 数据输出文件
DATA_OUTPUT_FILE = r'D:\Code\d2l-zh\my-new-site\day_log_data.json'
# 临时解压目录
TEMP_EXTRACT_DIR = r'D:\Code\d2l-zh\my-new-site\script\temp_extract'

# ================================================

class DiaryExtractor:
    def __init__(self):
        self.excel_dir = Path(EXCEL_DIR)
        self.images_dir = Path(IMAGES_OUTPUT_DIR)
        self.data_file = Path(DATA_OUTPUT_FILE)
        self.temp_dir = Path(TEMP_EXTRACT_DIR)

        # 创建必要的目录
        self.images_dir.mkdir(parents=True, exist_ok=True)

        # 加载现有数据
        self.existing_data = self.load_existing_data()
        self.existing_dates = set()

        if self.existing_data and 'rows' in self.existing_data:
            for row in self.existing_data['rows']:
                if '日期' in row and row['日期']:
                    self.existing_dates.add(row['日期'])

        print(f"✓ 已加载现有数据，包含 {len(self.existing_dates)} 条记录")

    def load_existing_data(self):
        """加载现有数据"""
        if self.data_file.exists():
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"⚠ 加载现有数据失败: {e}")
                return None
        return {'headers': [], 'rows': [], 'image_count': 0}

    def find_excel_files(self, pattern='*log*.xlsx'):
        """查找所有符合条件的 Excel 文件"""
        files = list(self.excel_dir.glob(pattern))
        print(f"\n找到 {len(files)} 个 Excel 文件:")
        for f in files:
            print(f"  - {f.name}")
        return files

    def extract_images_from_xlsx(self, excel_path):
        """从 xlsx 文件中提取图片"""
        print(f"\n正在提取图片: {excel_path.name}")

        # 清理临时目录
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)
        self.temp_dir.mkdir()

        # 解压 xlsx
        try:
            with zipfile.ZipFile(excel_path, 'r') as zip_ref:
                zip_ref.extractall(self.temp_dir)
        except Exception as e:
            print(f"✗ 解压失败: {e}")
            return {}

        # 查找图片
        image_mapping = {}
        media_dir = self.temp_dir / 'xl' / 'media'

        if media_dir.exists():
            for idx, img_file in enumerate(sorted(media_dir.iterdir()), 1):
                if img_file.suffix.lower() in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                    # 提取图片ID或使用序号
                    image_mapping[idx] = {
                        'source': img_file,
                        'name': img_file.name,
                        'ext': img_file.suffix
                    }

        print(f"  找到 {len(image_mapping)} 张图片")
        return image_mapping

    def analyze_excel_structure(self, excel_path):
        """分析 Excel 结构，找出图片所在行"""
        print(f"\n正在分析 Excel 结构: {excel_path.name}")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # 读取表头
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            headers.append(str(cell_value) if cell_value else f"列{col}")

        print(f"  列数: {len(headers)}")
        print(f"  行数: {ws.max_row}")

        # 找出图片列（通常是第4列）
        image_col_idx = None
        for col_idx, header in enumerate(headers, 1):
            if '图片' in header or '图' in header or 'image' in header.lower():
                image_col_idx = col_idx
                print(f"  图片列: 第{col_idx}列 ({header})")
                break

        # 找出包含图片标记的行
        image_rows = []
        if image_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row_idx, image_col_idx).value
                if cell_value and isinstance(cell_value, str) and 'DISPIMG' in cell_value:
                    image_rows.append(row_idx)

        print(f"  包含图片的行: {len(image_rows)} 行")

        return {
            'headers': headers,
            'image_col_idx': image_col_idx,
            'image_rows': image_rows,
            'max_row': ws.max_row
        }

    def process_excel_file(self, excel_path):
        """处理单个 Excel 文件"""
        print(f"\n{'='*60}")
        print(f"处理文件: {excel_path.name}")
        print(f"{'='*60}")

        # 提取图片
        image_mapping = self.extract_images_from_xlsx(excel_path)
        if not image_mapping:
            print("  ⚠ 未找到图片，跳过")
            return []

        # 分析结构
        structure = self.analyze_excel_structure(excel_path)

        # 读取数据
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        new_rows = []
        image_idx = 1

        for row_idx in range(2, structure['max_row'] + 1):
            row_data = {}
            has_data = False

            for col_idx in range(1, len(structure['headers']) + 1):
                cell_value = ws.cell(row_idx, col_idx).value
                header = structure['headers'][col_idx - 1]

                # 处理日期
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime('%Y-%m-%d')

                # 处理图片列
                if col_idx == structure['image_col_idx']:
                    # 检查这一行是否有图片
                    if row_idx in structure['image_rows']:
                        # 保存图片
                        if image_idx in image_mapping:
                            img_info = image_mapping[image_idx]
                            img_name = f"{excel_path.stem}_row{row_idx}{img_info['ext']}"
                            img_path = self.images_dir / img_name

                            # 复制图片
                            shutil.copy2(img_info['source'], img_path)
                            print(f"  ✓ 第{row_idx}行: {img_name}")

                            row_data[header] = img_name
                            has_data = True
                            image_idx += 1
                        else:
                            row_data[header] = None
                    else:
                        row_data[header] = None
                        if cell_value:
                            row_data[header] = str(cell_value)
                else:
                    if cell_value is not None:
                        row_data[header] = cell_value
                        has_data = True

            # 检查是否为新增数据
            if has_data and '日期' in row_data:
                if row_data['日期'] not in self.existing_dates:
                    new_rows.append(row_data)
                    self.existing_dates.add(row_data['日期'])

        # 清理临时文件
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

        print(f"\n✓ 新增 {len(new_rows)} 条记录")
        return new_rows, structure['headers']

    def save_data(self, new_rows, headers):
        """保存数据"""
        # 合并数据
        if 'rows' not in self.existing_data:
            self.existing_data['rows'] = []

        # 添加新行
        self.existing_data['rows'].extend(new_rows)

        # 按日期排序
        self.existing_data['rows'].sort(key=lambda x: x.get('日期', ''), reverse=True)

        # 更新表头
        self.existing_data['headers'] = headers

        # 统计图片
        image_count = sum(1 for row in self.existing_data['rows']
                         if any('图片' in k and v and isinstance(v, str) and v.endswith(('.png', '.jpg', '.jpeg'))
                               for k, v in row.items()))
        self.existing_data['image_count'] = image_count

        # 保存到文件
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(self.existing_data, f, ensure_ascii=False, indent=2)

        print(f"\n✓ 数据已保存到: {self.data_file}")
        print(f"✓ 总记录数: {len(self.existing_data['rows'])}")
        print(f"✓ 图片数量: {image_count}")

    def run(self, file_pattern='*log*.xlsx'):
        """运行提取流程"""
        print("\n" + "="*60)
        print("       Excel 日记增量提取工具")
        print("="*60)

        # 查找文件
        excel_files = self.find_excel_files(file_pattern)
        if not excel_files:
            print("\n未找到任何 Excel 文件")
            return

        # 处理每个文件
        all_new_rows = []
        headers = []

        for excel_file in excel_files:
            new_rows, file_headers = self.process_excel_file(excel_file)
            if new_rows:
                all_new_rows.extend(new_rows)
            if file_headers and not headers:
                headers = file_headers

        # 保存数据
        if all_new_rows:
            self.save_data(all_new_rows, headers)
            print(f"\n✓ 本次新增 {len(all_new_rows)} 条记录")
        else:
            print("\n⚠ 没有新增记录")

        print("\n" + "="*60)
        print("处理完成！")
        print("="*60)


if __name__ == "__main__":
    extractor = DiaryExtractor()

    # 处理所有包含 "log" 的 xlsx 文件
    # 也可以指定具体文件，如: 'day_log.xlsx'
    extractor.run('*log*.xlsx')
